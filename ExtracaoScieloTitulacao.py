#import selenium
#import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.relative_locator import locate_with
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
#from selenium.webdriver.support.ui import Select
from nltk.tokenize import RegexpTokenizer
import openpyxl
#from openpyxl.styles import PatternFill
import os
from time import sleep

#VARIÁVEIS PARA ALTERAÇÃO PELO USUÁRIO
diretorio_trabalho_pycharm = '/Users/Marcella/Downloads/UFMG/ExtracaoSCIELO'
nome_planilha = 'parquet.xlsx'
nome_folha = 'Planilha1'
caminho_webDriver = '/Users/Marcella/PycharmProjects/webscrapper101/chromedriver'



# ABRINDO O EXCEL
# muda o diretório que o pycharm está trabalhando
os.chdir(diretorio_trabalho_pycharm)

# inicializando driver
caminho = Service(caminho_webDriver)
navegador = webdriver.Chrome(service=caminho)

# Abre Planilha
planilha = openpyxl.load_workbook(nome_planilha)

# Escolha a folha
folha1 = planilha[nome_folha]

# cria uma lista para colocar os links
links = []

# popula a lista vazia de links
for linha in folha1:
    url = linha[12].value
    links.append(url)

# Print lista de links
print(links)

adjetivosDoutor = ['doutor', 'doutorado', 'doutorae', 'doutora', 'titular']
adjetivosMestre = ['mestre', 'mestrado', 'mestra', 'doutorando', 'doutoranda']

adjetivosGraduacao = ['mestranda', 'mestrando', 'graduando', 'graduanda', 'bacharel', 'bacharela', 'bacharelado',
                      'bachalerato']

for n in range(len(links)):
    navegador.get(links[n])
    sleep(3)
    try:
        caixaAutoria = navegador.find_element(By.CSS_SELECTOR,
                                              '#standalonearticle > section > div > div > div.contribGroup > a:nth-child(1)')
        caixaAutoria.click()
    except NoSuchElementException or ElementClickInterceptedException:
        pass

    sleep(5)

    try:
        instituicao = navegador.find_element(By.CSS_SELECTOR,
                                             '#ModalTutors > div > div > div.modal-body > div.info > div:nth-child(1) > div:nth-child(3) > span:nth-child(2)')
        celulaInstituicao = folha1.cell(row=n + 1, column=5)
        celulaInstituicao.value = instituicao.get_attribute('textContent').upper()
    except NoSuchElementException:
        try:
            instituicao = navegador.find_element(By.CSS_SELECTOR,
                                                 '#ModalTutors > div > div > div.modal-body > div.info > div > div:nth-child(4) > span:nth-child(2)')
            celulaInstituicao = folha1.cell(row=n + 1, column=5)
            celulaInstituicao.value = instituicao.get_attribute('textContent').upper()
        except NoSuchElementException:
            try:
                instituicao = navegador.find_element(By.CSS_SELECTOR,
                                                     '#ModalTutors > div > div > div.modal-body > div.info > div:nth-child(1) > div:nth-child(5) > span:nth-child(2)')
                celulaInstituicao = folha1.cell(row=n + 1, column=5)
                celulaInstituicao.value = instituicao.get_attribute('textContent').upper()
            except NoSuchElementException:
                try:
                    instituicao = navegador.find_element(By.CSS_SELECTOR,
                                                         '#ModalTutors > div > div > div.modal-body > div.info > div:nth-child(1) > div:nth-child(10) > span:nth-child(1)')
                    celulaInstituicao = folha1.cell(row=n + 1, column=5)
                    celulaInstituicao.value = instituicao.get_attribute('textContent').upper()
                except NoSuchElementException:
                    try:
                        instituicao = navegador.find_element(By.CSS_SELECTOR,
                                                             '#ModalTutors > div > div > div.modal-body > div.info > div:nth-child(1) > div:nth-child(11) > span:nth-child(1)')
                        celulaInstituicao = folha1.cell(row=n + 1, column=5)
                        celulaInstituicao.value = instituicao.get_attribute('textContent').upper()
                    except NoSuchElementException:
                        pass


    try:
        localidade = navegador.find_element(By.CSS_SELECTOR,
                                            '#ModalTutors > div > div > div.modal-body > div.info > div:nth-child(1) > div:nth-child(3) > span:nth-child(5)')
        textoLocalidade = localidade.get_attribute('textContent').upper()

        listaLocalidade = textoLocalidade.split(',')

        celulaPais = folha1.cell(row=n + 1, column=10)
        try:
            celulaPais.value = listaLocalidade[-1]
        except IndexError:
            pass

        celulaEstado = folha1.cell(row=n + 1, column=11)
        try:
            celulaEstado.value = listaLocalidade[-2]
        except IndexError:
            pass

        celulaCidade = folha1.cell(row=n + 1, column=12)
        try:
            celulaCidade.value = listaLocalidade[-3]
        except IndexError:
            pass

    except NoSuchElementException:
        try:
            localidade = navegador.find_element(By.CSS_SELECTOR,
                                                '#ModalTutors > div > div > div.modal-body > div.info > div > div:nth-child(4) > span:nth-child(5)')
            textoLocalidade = localidade.get_attribute('textContent').upper()

            listaLocalidade = textoLocalidade.split(',')

            celulaPais = folha1.cell(row=n + 1, column=10)
            try:
                celulaPais.value = listaLocalidade[-1]
            except IndexError:
                pass

            celulaEstado = folha1.cell(row=n + 1, column=11)
            try:
                celulaEstado.value = listaLocalidade[-2]
            except IndexError:
                pass

            celulaCidade = folha1.cell(row=n + 1, column=12)
            try:
                celulaCidade.value = listaLocalidade[-3]
            except IndexError:
                pass
        except NoSuchElementException:
            try:
                localidade = navegador.find_element(By.CSS_SELECTOR,
                                                    '#ModalTutors > div > div > div.modal-body > div.info > div:nth-child(1) > div:nth-child(10) > span:nth-child(5)')
                textoLocalidade = localidade.get_attribute('textContent').upper()

                listaLocalidade = textoLocalidade.split(',')

                celulaPais = folha1.cell(row=n + 1, column=10)
                try:
                    celulaPais.value = listaLocalidade[-1]
                except IndexError:
                    pass

                celulaEstado = folha1.cell(row=n + 1, column=11)
                try:
                    celulaEstado.value = listaLocalidade[-2]
                except IndexError:
                    pass

                celulaCidade = folha1.cell(row=n + 1, column=12)
                try:
                    celulaCidade.value = listaLocalidade[-3]
                except IndexError:
                    pass
            except NoSuchElementException:
                pass

    try:
        curriculo1autor = navegador.find_element(By.CSS_SELECTOR,
                                                 '#ModalTutors > div > div > div.modal-body > div:nth-child(2) > ul > li > div')
        curriculo1autorTokenizado = RegexpTokenizer(r'\w+').tokenize(curriculo1autor.text.lower())

        celulaFormacao1Autor = folha1.cell(row=n + 1, column=4)

        if any(palavra in adjetivosDoutor for palavra in curriculo1autorTokenizado):
            celulaFormacao1Autor.value = 'Doutorado'
        elif any(palavra in adjetivosMestre for palavra in curriculo1autorTokenizado):
            celulaFormacao1Autor.value = 'Mestrado'
        elif any(palavra in adjetivosGraduacao for palavra in curriculo1autorTokenizado):
            celulaFormacao1Autor.value = 'Graduação'
        else:
            celulaFormacao1Autor.value = ' '
    except NoSuchElementException:
        try:
            curriculo1autor = navegador.find_element(By.CSS_SELECTOR,
                                                     '#ModalTutors > div > div > div.modal-body > div:nth-child(3) > ul > li > div')
            curriculo1autorTokenizado = RegexpTokenizer(r'\w+').tokenize(curriculo1autor.text.lower())

            celulaFormacao1Autor = folha1.cell(row=n + 1, column=4)

            if any(palavra in adjetivosDoutor for palavra in curriculo1autorTokenizado):
                celulaFormacao1Autor.value = 'Doutorado'
            elif any(palavra in adjetivosMestre for palavra in curriculo1autorTokenizado):
                celulaFormacao1Autor.value = 'Mestrado'
            elif any(palavra in adjetivosGraduacao for palavra in curriculo1autorTokenizado):
                celulaFormacao1Autor.value = 'Graduação'
            else:
                celulaFormacao1Autor.value = ' '
        except NoSuchElementException:
            pass

    planilha.save(nome_planilha)
