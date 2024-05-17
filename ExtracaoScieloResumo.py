#import selenium
#import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.relative_locator import locate_with
from selenium.common.exceptions import NoSuchElementException
#from selenium.webdriver.support.ui import Select
#from nltk.tokenize import RegexpTokenizer
import openpyxl
#from openpyxl.styles import PatternFill
import os
from time import sleep

#ESSE CÓDIGO EXTRAI RESUMO PALAVRAS CHAVE MES PUBLICACAO E ANO PUBLICACAO

#VARIÁVEIS PARA ALTERAÇÃO PELO USUÁRIO
diretorio_trabalho_pycharm = '/Users/Marcella/Downloads/UFMG/ExtracaoSCIELO'
nome_planilha = 'parquet.xlsx'
nome_folha = 'Planilha1'
caminho_webDriver = '/Users/Marcella/PycharmProjects/webscrapper101/chromedriver'

# ABRINDO O EXCEL
# muda o diretório que o pycharm está trabalhando
os.chdir(diretorio_trabalho_pycharm)

# inicializando driver
caminho = Service(caminho_webDriver)
navegador = webdriver.Chrome(service=caminho)

# Abre Planilha
planilha = openpyxl.load_workbook(nome_planilha)

# Escolha a folha
folha1 = planilha[nome_folha]

# cria uma lista para colocar os links
links = []

# popula a lista vazia de links com o conteúdo da coluna 12
for linha in folha1:
    url = linha[12].value
    links.append(url)

# Print lista de links
print(links)

textoPalavraChave = 'palavras-chave'
textoResumo = 'resumo'

for n in range(len(links)):
    navegador.get(links[n])
    sleep(4)
    source = navegador.page_source.lower()

    try:
        textoResumo = navegador.find_element(
        locate_with(By.TAG_NAME, "p").below({By.XPATH: "//*[contains(text(),'Resumo')]"}))
        resumo = textoResumo.text
        celulaResumo = folha1.cell(row=n + 1, column=14)
        celulaResumo.value = resumo
    except NoSuchElementException:
        pass


    if textoPalavraChave in source:
        divsPalavraChave = navegador.find_elements(By.CSS_SELECTOR, 'p:nth-child(4)')
        if len(divsPalavraChave) > 0:
            palavrasChave = divsPalavraChave[0].text.lower()
            celulaPalavrasChave = folha1.cell(row=n + 1, column=15)
            celulaPalavrasChave.value = palavrasChave

    #encontra data de publicação
    try:
        dataPublicacao = navegador.find_element(By.CSS_SELECTOR,
                                                '#articleText > div:nth-child(6) > div > div > ul > li:nth-child(1)')

        celulaMesPublicacao = folha1.cell(row=n + 1, column=8)
        celulaMesPublicacao.value = dataPublicacao.text[-9:-5] # mês de publicação

        celulaAnoPublicacao = folha1.cell(row=n + 1, column=9)
        celulaAnoPublicacao.value = dataPublicacao.text[-5:] # ano de publicação
    except NoSuchElementException:
        try:
            dataPublicacao = navegador.find_element(By.CSS_SELECTOR,
                                                    '#articleText > div:nth-child(7) > div > div > ul > li:nth-child(1)')

            celulaMesPublicacao = folha1.cell(row=n + 1, column=8)
            celulaMesPublicacao.value = dataPublicacao.text[-9:-5]  # mês de publicação

            celulaAnoPublicacao = folha1.cell(row=n + 1, column=9)
            celulaAnoPublicacao.value = dataPublicacao.text[-5:]  # ano de publicação
        except NoSuchElementException:
            try:
                dataPublicacao = navegador.find_element(By.CSS_SELECTOR,
                                                        '#articleText > div:nth-child(9) > div > div > ul > li:nth-child(1)')

                celulaMesPublicacao = folha1.cell(row=n + 1, column=8)
                celulaMesPublicacao.value = dataPublicacao.text[-9:-5]  # mês de publicação

                celulaAnoPublicacao = folha1.cell(row=n + 1, column=9)
                celulaAnoPublicacao.value = dataPublicacao.text[-5:]  # ano de publicação
            except NoSuchElementException:
                try:
                    dataPublicacao = navegador.find_element(By.CSS_SELECTOR,
                                                            '#articleText > div:nth-child(10) > div > div > ul > li:nth-child(1)')

                    celulaMesPublicacao = folha1.cell(row=n + 1, column=8)
                    celulaMesPublicacao.value = dataPublicacao.text[-9:-5]  # mês de publicação

                    celulaAnoPublicacao = folha1.cell(row=n + 1, column=9)
                    celulaAnoPublicacao.value = dataPublicacao.text[-5:]  # ano de publicação
                except NoSuchElementException:
                    try:
                        dataPublicacao = navegador.find_element(By.CSS_SELECTOR,
                                                                '#articleText > div:nth-child(11) > div > div > ul > li:nth-child(1)')

                        celulaMesPublicacao = folha1.cell(row=n + 1, column=8)
                        celulaMesPublicacao.value = dataPublicacao.text[-9:-5]  # mês de publicação

                        celulaAnoPublicacao = folha1.cell(row=n + 1, column=9)
                        celulaAnoPublicacao.value = dataPublicacao.text[-5:]  # ano de publicação
                    except NoSuchElementException:
                        pass

    planilha.save(nome_planilha)

