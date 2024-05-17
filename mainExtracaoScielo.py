from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from nltk.tokenize import RegexpTokenizer
import openpyxl
import os
from time import sleep


def proxPagina(numPag):
    navegador.get(f'https://search.scielo.org/?lang=en&count=15&from={(15 * (numPag - 1))+1}&output=site&sort=&format=summary&fb=&page={numPag}&q=%22parquet%22')
    #navegador.get(f'https://search.scielo.org/?q=justi%C3%A7a+trabalhista&lang=en&count=15&from={(15 * (numPag - 1))+1}&output=site&sort=&format=summary&fb=&page={numPag}&filter%5Bin%5D%5B%5D=scl&q=justi%C3%A7a&lang=en')


#ESTE CÓDIGO EXTRAI TITULO, AUTORES, PRIMEIRO AUTOR, FONTE, LINKS

#VARIÁVEIS PARA ALTERAÇÃO PELO USUÁRIO
diretorio_trabalho_pycharm = '/Users/Marcella/Downloads/UFMG/ExtracaoSCIELO'
nome_planilha = 'parquet.xlsx'
nome_folha = 'Planilha1'
caminho_webDriver = '/Users/Marcella/PycharmProjects/webscrapper101/chromedriver'
pagina_inicial_pesquisa = 'https://search.scielo.org/?lang=en&count=15&from=0&output=site&sort=&format=summary&fb=&page=1&q=%22parquet%22'
num_paginas = 2

#COMEÇANDO O CÓDIGO

# ABRINDO O EXCEL
# muda o diretório que o pycharm está trabalhando
os.chdir(diretorio_trabalho_pycharm)

# abre (ou carrega) a planilha no excel
planilha = openpyxl.load_workbook(nome_planilha)

# pega a primeira folha da planilha. (aquelas abas debaixo). O caminho de antes era get sheet by name
folha1 = planilha[nome_folha]

# Abre navegador já na página de resultado de uma pesquisa na scielo
caminho = Service(caminho_webDriver)
navegador = webdriver.Chrome(service=caminho)
navegador.get(pagina_inicial_pesquisa)

# esse link tem que ser dá página de busca já nos resultados. TEM QUE MUDAR NA FUNÇÃO TAMBÉM
sleep(7)  # tempo para carregar a página

listaTitlesErrados = ['Selected filters\nCLEAN', 'Collection', 'Journal\n+ OPTIONS', 'Language', 'Publication Year\n+ OPTIONS',
               'SciELO Thematic Areas\n+ OPTIONS', 'WoS Subject Categories\n+ OPTIONS',
               'WoS Citation Index\n+ OPTIONS', 'Citables and non citables', 'Type of Literature', 'Type of Literature\n+ OPTIONS',
               'Citables and non citables\n+ OPTIONS', 'Language\n+ OPTIONS', 'WoS Citation Index']

listaNomesFemininos = [
    "Amanda", "Ana", "Antônia", "Aparecida", "Beatriz", "Bianca", "Camila", "Carolina", "Catarina", "Carmen", "Cecília",
    "Cristiana", "Cristiane", "Carla", "Ilma", "Eloá", "Marina", "Marcela", "Marcella", "Mônica", "Monica",
    "Clara", "Diana", "Daniela", "Elisabete", "Eloísa", "Estela", "Fernanda", "Fábia", "Gabriela", "Giselle",
    "Helena", "Isabel", "Joana", "Júlia", "Juliana", "Laura", "Lara", "Leonor", "Liliane", "Lígia", "Lisandra",
    "Lorena", "Luciana", "Lúcia", "Luísa", "Mariana", "Marjorie", "Manuela", "Maria", "Marianna", "Mirella", "Natália", "Rayane"
    "Olívia", "Patrícia", "Rafaela", "Regina", "Renata", "Rita", "Selma", "Sofia", "Silvia", "Sílvia", "Tatiana", "Valentina",
    "Vanessa", "Victoria", "Victória", "Vitória", "Vitoria", "Andreza", "Andrezza", "Letícia", "Leticia", "Nayara", "Adriana",
    "Larissa", "Bárbara", "Barbara", "Marília", "Fabiana", "Camila", "Gisleuda", "Susana", "Bruna", "Cynthia", "Cinthia", "Cintia",
    "Cyntia", "Juliane", "Natalia", "Nathalia", "Natalie", "Aline", "Caroline", "Jordana", "Janaina", "Janaína", "Flávia",
    "Liz", "Monique"
]

listaNomesMasculinos = [
    "Adriano", "Alexandre", "André", "Artur", "Arthur", "Bernardo", "Carlos", "Daniel", "Eduardo", "Felipe", "Gustavo", "Henrique",
    "Isaac", "João", "Lucas", "Marcos", "Nuno", "Otávio", "Paulo", "Rafael", "Samuel", "Tiago", "Vítor",
    "Antônio", "Benjamin", "César", "Diego", "Emanuel", "Fernando", "Gabriel", "Hugo", "Igor", "Juliano",
    "Kleber", "Lorenzo", "Miguel", "Nathan", "Nivaldo", "Oscar", "Pedro", "Ricardo", "Rogerio", "Sérgio", "Sergio",
    "Sebastião", "Thiago", "Ulisses", "Valentim", "ßßMario", "Marcio", "Fábio", "Fabio", "Francisco", "Sandro", "Luiz",
    "Luis", "Luíz", "José", "Thiago", "Felipe", "Filipe", "Pablo", "Gilberto", "Bruno", "Matheus", "Mateus", "Michel",
    "Cleber", "Airton", "Luís", "Luiz", "Nicholas", "Márcio"
]

for x in range(0, (num_paginas + 1)): #num de paginas + 1
    listaTitulosCertos = []
    listaTudoTitle = navegador.find_elements(By.CLASS_NAME, 'title')

    for n in range(len(listaTudoTitle)):
        if listaTudoTitle[n].text not in listaTitlesErrados:
            listaTitulosCertos.append(listaTudoTitle[n].text) #popula a lista de títulos certos

    #encontra autores
    listaAutores = navegador.find_elements(By.CSS_SELECTOR, 'div.line.authors')


    #encontra links
    listaLinks = navegador.find_elements(By.CSS_SELECTOR, 'div.col-md-11.col-sm-10.col-xs-11 > div:nth-child(1) > a')

    # encontra fonte
    listaFontes = navegador.find_elements(By.CSS_SELECTOR, 'div.line.source')

    for n in range(len(listaTitulosCertos)):
        celulaTitulo = folha1.cell(row=(n+1)+(x*15), column=1)
        celulaTitulo.value = listaTitulosCertos[n].upper()

        celulaAutores = folha1.cell(row=(n+1)+(x*15), column=2)

        try:
            celulaAutores.value = listaAutores[n].text.title()

            # achando o nome do primeiro autor
            nomesAutores = listaAutores[n].text.title()
            nome1autor = nomesAutores.split(';')[0]
            nome1autorsplit = nome1autor.split(',')
            nome1autorJoin = f'{nome1autorsplit[1]}{nome1autorsplit[0]}'
            nome1autorTokenizado = RegexpTokenizer(r'\w+').tokenize(nome1autorJoin.title())

            # colocando o nome do 1º autor na célula
            celula1Autor = folha1.cell(row=(n + 1) + (x * 15), column=3)
            celula1Autor.value = ' '.join(nome1autorTokenizado)

            celulaGenero1Autor = folha1.cell(row=(n + 1) + (x * 15), column=6)
            if nome1autorTokenizado[0] in listaNomesFemininos:
                celulaGenero1Autor.value = 'Feminino'
            elif nome1autorTokenizado[0] in listaNomesMasculinos:
                celulaGenero1Autor.value = 'Masculino'
            else:
                celulaGenero1Autor.value = ' '
        except IndexError:
            pass

        celulaFonte = folha1.cell(row=(n + 1) + (x * 15), column=7)
        try:
            celulaFonte.value = listaFontes[n].text
        except IndexError:
            pass

        celulaLink = folha1.cell(row=(n+1)+(x*15), column=13)

        try:
            celulaLink.value = listaLinks[n].get_attribute('href')
        except IndexError:
            pass

    planilha.save(nome_planilha)

    proxPagina(x+1)







